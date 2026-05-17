#!/usr/bin/env python3
"""
Palfinger varilni izvori - izvoz v XLSX
Uporaba: python3 palfinger_export.py podatki.json palfinger_glavni.xlsx [izhod.xlsx]
"""
import sys, json, shutil, datetime
from openpyxl import load_workbook

def to_num(v):
    if v is None or v == '' or v == '/' or v == 'N/A':
        return v
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return v

def parse_date(s):
    if not s: return None
    try: return datetime.datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except: return s

def update_palfinger(input_path, output_path, header, blocks):
    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['Osnova']

    # Header
    ws['B1']  = parse_date(header.get('datum'))
    ws['E1']  = to_num(header.get('zacetna_st'))
    ws['B2']  = header.get('narocnik', '')
    ws['B3']  = header.get('naslov', '')
    ws['B4']  = header.get('narocilo', '')
    ws['B5']  = to_num(header.get('DN'))
    ws['B6']  = header.get('porocilo_st', '')
    ws['B7']  = header.get('kraj_datum', '')
    ws['B8']  = to_num(header.get('temperatura'))
    ws['B9']  = to_num(header.get('vlaga'))
    ws['B10'] = parse_date(header.get('datum_nalepke'))

    # Device blocks - offsets from start_row (1-indexed)
    str_fields = {'tov_st', 'inv_st', 'proizvajalec', 'postopek', 'tip'}
    field_offsets = [
        (0, 'seq'), (2, 'tip'), (3, 'postopek'), (4, 'tov_st'), (5, 'inv_st'),
        (6, 'proizvajalec'), (7, 'presek'), (8, 'padec'),
        (9, 'vhodni_25'), (10, 'varilni_25'), (11, 'vhodni_5'),
        (12, 'U0x'), (13, 'U0s'), (14, 'U21x'), (15, 'U21s'),
        (16, 'U22x'), (17, 'U22s'),
        (18, 'I2x1'), (19, 'I2s1'), (20, 'I2x2'), (21, 'I2s2'),
        (22, 'I2x3'), (23, 'I2s3'), (24, 'Imax'),
    ]

    for block in blocks:
        sr = block['start_row']
        for offset, field in field_offsets:
            raw = block.get(field)
            if field == 'seq':
                val = int(raw) if raw is not None else None
            elif field in str_fields:
                val = str(raw) if raw is not None else None
            else:
                val = to_num(raw)
            ws.cell(row=sr + offset, column=2).value = val

        # Second U22 values in column C
        ws.cell(row=sr + 16, column=3).value = to_num(block.get('U22x2'))
        ws.cell(row=sr + 17, column=3).value = to_num(block.get('U22s2'))

    wb.save(output_path)
    print(f"✓ Shranjeno: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Uporaba: python3 palfinger_export.py podatki.json palfinger_glavni.xlsx [izhod.xlsx]")
        sys.exit(1)

    json_path   = sys.argv[1]
    input_xlsx  = sys.argv[2]
    output_xlsx = sys.argv[3] if len(sys.argv) > 3 else 'palfinger_izhod.xlsx'

    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)

    update_palfinger(input_xlsx, output_xlsx, data['header'], data['blocks'])
